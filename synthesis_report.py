#!/usr/bin/python
import sys
sys.path.append("./openpyxl")

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl.style import Color

import re
import os
import os.path
from decimal import *
from optparse import OptionParser
import glob
from datetime import datetime

parser = OptionParser()

parser.add_option("-d", "--directory", action="store", type="string", dest="dirName",
        help="the directory of the report folder", metavar="DIR")
parser.add_option("-t", "--title", action="store", type="string", dest="title",
        help="the tittle of the record", metavar="STRING")
parser.add_option("-r", "--replace", action="store_true", dest="replace", default=False,
        help="overwrite the file", )
parser.add_option("-f", "--file", action="store", dest="destFile", default="report.xls",
        help="output file", metavar="FILE")

(options, args) = parser.parse_args()

if not (options.dirName):
    parser.print_help()
    exit()

dirName = options.dirName
title = options.title
report = options.destFile
overwrite = options.replace


if not (os.path.exists(dirName)):
    error = "folder is %s not found\n"%(dirName)
    exit()

creat = False
if not (os.path.isfile(report) ):
    info = "file %s not found, create a new one"%(report)
    creat = True
    print info

if creat:
    wb = Workbook()
else:
    wb = load_workbook(report)

ws = wb.get_sheet_by_name(r'report')
if not ws:
    ws = wb.create_sheet();
    ws.title = "report";
    firstBlank = 0;

firstBlank = ws.get_highest_column() - 1
#print firstBlank

if not (firstBlank == 0):
    d = dict()
    for rowIndex in xrange(0, ws.get_highest_row()):
        value = ws.cell(row = rowIndex, column = 0).value
        d[value] = rowIndex

i = 2
j = firstBlank
if not title:
    title = dirName


ws.cell(row = 0, column = j + 1).value = str(datetime.now())
ws.cell(row = 1, column = j + 1).value = title
ws.cell(row = 1, column = j + 1).style.font.color.index = Color.GREEN

regExp = "^\s*Total cell area:\s+(?P<value>[0-9]+\.[0-9]+)\s*$"
for fileName in glob.glob(os.path.join(dirName, '*.area')):
    basename = os.path.basename(fileName)
    basename = os.path.splitext(basename)[0]
    file = open(fileName, 'r+');
    for line in file:
        m = re.search(regExp, line)
        if m:
            value = m.group('value')
            if firstBlank == 0:
                ws.cell(row = i, column = j).value =basename
                ws.cell(row = i, column = j + 1).value = value
            else:
                if not d.has_key(basename):
                    last = ws.get_highest_row()
                    ws.cell(row = last, column = 0).value =basename
                    ws.cell(row = last, column = j + 1).value = value
                else:
                    index = d[basename]
                    ws.cell(row = index, column = j + 1).value = value
            break
    i = i +1   

col = get_column_letter(j + 2)
ws.column_dimensions["A"].width = 20
ws.column_dimensions[col].width = 30
wb.save(report)



